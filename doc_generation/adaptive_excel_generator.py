# ============================================================
# YantraAI - Adaptive Excel Generator
# JSON -> Microsoft Excel (.xlsx)
# ============================================================

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# HELPERS
# ============================================================

def safe_filename(name):
    name = str(name or "document")

    name = re.sub(
        r'[<>:"/\\|?*]',
        "_",
        name
    )

    return name.strip().replace(" ", "_")


def humanize_key(key):
    key = str(key)
    key = key.replace("_", " ")

    key = re.sub(
        r"([a-z])([A-Z])",
        r"\1 \2",
        key
    )

    return key.title()


def format_value(value):

    if value is None:
        return ""

    if isinstance(value, bool):
        return "Yes" if value else "No"

    if isinstance(value, (dict, list)):
        return json.dumps(
            value,
            ensure_ascii=False
        )

    return value


# ============================================================
# ADAPTIVE EXCEL GENERATOR
# ============================================================

class AdaptiveExcelGenerator:

    def __init__(self, data):

        self.data = data

        self.workbook = Workbook()

        # Remove default sheet
        default_sheet = (
            self.workbook.active
        )

        self.workbook.remove(
            default_sheet
        )

        formatting = (
            data.get("formatting", {})
        )

        if not formatting:

            content = data.get(
                "content",
                {}
            )

            if isinstance(
                content,
                dict
            ):

                formatting = content.get(
                    "formatting",
                    {}
                )

        if not isinstance(
            formatting,
            dict
        ):
            formatting = {}

        self.header_bold = (
            formatting.get(
                "bold",
                True
            )
        )

        self.font_name = (
            formatting.get(
                "font",
                "Arial"
            )
        )

        self.font_size = (
            self._font_size(
                formatting.get(
                    "size",
                    11
                )
            )

        )

    # ========================================================
    # FONT SIZE
    # ========================================================

    def _font_size(self, value):

        try:

            match = re.search(
                r"\d+(?:\.\d+)?",
                str(value)
            )

            if match:

                return float(
                    match.group()
                )

        except Exception:
            pass

        return 11

    # ========================================================
    # CREATE SHEET
    # ========================================================

    def create_sheet(
        self,
        name
    ):

        name = str(name)

        # Excel sheet name restrictions
        name = re.sub(
            r'[\\/*?:\[\]]',
            "_",
            name
        )

        name = name[:31]

        if not name:
            name = "Sheet1"

        return self.workbook.create_sheet(
            title=name
        )

    # ========================================================
    # STYLE HEADER
    # ========================================================

    def style_header(
        self,
        cells
    ):

        for cell in cells:

            cell.font = Font(
                name=self.font_name,
                size=self.font_size,
                bold=self.header_bold
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

    # ========================================================
    # STYLE NORMAL CELL
    # ========================================================

    def style_cell(
        self,
        cell
    ):

        cell.font = Font(
            name=self.font_name,
            size=self.font_size
        )

        cell.alignment = Alignment(
            vertical="center",
            wrap_text=True
        )

    # ========================================================
    # AUTO WIDTH
    # ========================================================

    def auto_width(
        self,
        worksheet
    ):

        for column in worksheet.columns:

            max_length = 0

            column_letter = (
                get_column_letter(
                    column[0].column
                )
            )

            for cell in column:

                try:

                    length = len(
                        str(cell.value)
                    )

                    if length > max_length:
                        max_length = length

                except Exception:
                    pass

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max(max_length + 2, 12),
                50
            )

    # ========================================================
    # ADD TABLE
    # ========================================================

    def add_table(
        self,
        worksheet,
        columns,
        rows,
        start_row=1,
        start_col=1,
        title=None
    ):

        current_row = start_row

        # ----------------------------------------------------
        # TABLE TITLE
        # ----------------------------------------------------

        if title:

            cell = worksheet.cell(
                row=current_row,
                column=start_col
            )

            cell.value = str(title)

            cell.font = Font(
                name=self.font_name,
                size=self.font_size + 2,
                bold=True
            )

            current_row += 2

        # ----------------------------------------------------
        # HEADER
        # ----------------------------------------------------

        for index, column in enumerate(
            columns
        ):

            cell = worksheet.cell(
                row=current_row,
                column=start_col + index
            )

            cell.value = (
                humanize_key(column)
            )

        header_cells = [
            worksheet.cell(
                row=current_row,
                column=start_col + i
            )
            for i in range(
                len(columns)
            )
        ]

        self.style_header(
            header_cells
        )

        current_row += 1

        # ----------------------------------------------------
        # DATA
        # ----------------------------------------------------

        for row in rows:

            if isinstance(
                row,
                dict
            ):

                values = [
                    row.get(
                        column,
                        ""
                    )
                    for column in columns
                ]

            else:

                values = list(row)

            for index in range(
                len(columns)
            ):

                value = (
                    values[index]
                    if index < len(values)
                    else ""
                )

                cell = worksheet.cell(
                    row=current_row,
                    column=start_col + index
                )

                cell.value = (
                    format_value(value)
                )

                self.style_cell(
                    cell
                )

            current_row += 1

        # ----------------------------------------------------
        # BORDER
        # ----------------------------------------------------

        thin = Side(
            style="thin"
        )

        border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

        end_row = (
            current_row - 1
        )

        for row in worksheet.iter_rows(
            min_row=start_row,
            max_row=end_row,
            min_col=start_col,
            max_col=(
                start_col
                + len(columns)
                - 1
            )
        ):

            for cell in row:

                cell.border = border

        return current_row

    # ========================================================
    # DICTIONARY
    # ========================================================

    def render_dict(
        self,
        worksheet,
        data,
        start_row=1
    ):

        row = start_row

        for key, value in data.items():

            if key == "formatting":
                continue

            # Nested dictionary
            if isinstance(
                value,
                dict
            ):

                cell = worksheet.cell(
                    row=row,
                    column=1
                )

                cell.value = (
                    humanize_key(key)
                )

                cell.font = Font(
                    name=self.font_name,
                    size=self.font_size + 1,
                    bold=True
                )

                row += 1

                row = self.render_dict(
                    worksheet,
                    value,
                    row
                )

            # List
            elif isinstance(
                value,
                list
            ):

                if not value:
                    continue

                # List of dictionaries
                if all(
                    isinstance(
                        item,
                        dict
                    )
                    for item in value
                ):

                    columns = []

                    for item in value:

                        for k in item:

                            if k not in columns:

                                columns.append(k)

                    row = self.add_table(
                        worksheet,
                        columns,
                        value,
                        row,
                        1,
                        humanize_key(key)
                    )

                else:

                    worksheet.cell(
                        row=row,
                        column=1
                    ).value = (
                        humanize_key(key)
                    )

                    row += 1

                    for item in value:

                        worksheet.cell(
                            row=row,
                            column=1
                        ).value = (
                            format_value(item)
                        )

                        row += 1

            # Primitive value
            else:

                worksheet.cell(
                    row=row,
                    column=1
                ).value = (
                    humanize_key(key)
                )

                worksheet.cell(
                    row=row,
                    column=2
                ).value = (
                    format_value(value)
                )

                row += 1

        return row

    # ========================================================
    # SECTIONS
    # ========================================================

    def render_sections(
        self,
        worksheet,
        sections
    ):

        row = 1

        for section in sections:

            if not isinstance(
                section,
                dict
            ):
                continue

            title = (
                section.get(
                    "section_title"
                )
                or section.get(
                    "title"
                )
                or section.get(
                    "name"
                )
                or "Section"
            )

            cell = worksheet.cell(
                row=row,
                column=1
            )

            cell.value = str(title)

            cell.font = Font(
                name=self.font_name,
                size=self.font_size + 2,
                bold=True
            )

            row += 2

            content = section.get(
                "content"
            )

            if isinstance(
                content,
                dict
            ):

                row = self.render_dict(
                    worksheet,
                    content,
                    row
                )

            elif isinstance(
                content,
                list
            ):

                if all(
                    isinstance(
                        item,
                        dict
                    )
                    for item in content
                ):

                    columns = []

                    for item in content:

                        for key in item:

                            if key not in columns:

                                columns.append(
                                    key
                                )

                    row = self.add_table(
                        worksheet,
                        columns,
                        content,
                        row,
                        1
                    )

                else:

                    for item in content:

                        worksheet.cell(
                            row=row,
                            column=1
                        ).value = (
                            format_value(item)
                        )

                        row += 1

            elif content is not None:

                worksheet.cell(
                    row=row,
                    column=1
                ).value = (
                    format_value(content)
                )

                row += 1

            row += 1

        return row

    # ========================================================
    # TABLES
    # ========================================================

    def render_tables(
        self,
        worksheet,
        tables
    ):

        row = 1

        for table in tables:

            if not isinstance(
                table,
                dict
            ):
                continue

            columns = (
                table.get("columns")
                or table.get("headers")
                or []
            )

            rows = (
                table.get("rows")
                or table.get("data")
                or []
            )

            title = (
                table.get("table_title")
                or table.get("title")
                or table.get("name")
            )

            if columns:

                row = self.add_table(
                    worksheet,
                    columns,
                    rows,
                    row,
                    1,
                    title
                )

                row += 1

        return row

    # ========================================================
    # MAIN RENDER
    # ========================================================

    def render(self):

        content = self.data.get(
            "content",
            {}
        )

        # ----------------------------------------------------
        # If explicit tables/sections exist
        # ----------------------------------------------------

        if isinstance(
            content,
            dict
        ):

            sections = content.get(
                "sections"
            )

            tables = content.get(
                "tables"
            )

            # Create dedicated sheets
            if sections:

                for index, section in enumerate(
                    sections,
                    1
                ):

                    if isinstance(
                        section,
                        dict
                    ):

                        sheet_name = (
                            section.get(
                                "section_title"
                            )
                            or section.get(
                                "title"
                            )
                            or f"Section_{index}"
                        )

                    else:

                        sheet_name = (
                            f"Section_{index}"
                        )

                    worksheet = (
                        self.create_sheet(
                            sheet_name
                        )
                    )

                    self.render_sections(
                        worksheet,
                        [section]
                    )

                    self.auto_width(
                        worksheet
                    )

            if tables:

                worksheet = (
                    self.create_sheet(
                        "Tables"
                    )
                )

                self.render_tables(
                    worksheet,
                    tables
                )

                self.auto_width(
                    worksheet
                )

            # ------------------------------------------------
            # Other dynamic content
            # ------------------------------------------------

            for key, value in content.items():

                if key in {
                    "sections",
                    "tables",
                    "formatting"
                }:
                    continue

                worksheet = (
                    self.create_sheet(
                        key
                    )
                )

                if isinstance(
                    value,
                    dict
                ):

                    self.render_dict(
                        worksheet,
                        value
                    )

                elif isinstance(
                    value,
                    list
                ):

                    if all(
                        isinstance(
                            item,
                            dict
                        )
                        for item in value
                    ):

                        columns = []

                        for item in value:

                            for k in item:

                                if k not in columns:

                                    columns.append(
                                        k
                                    )

                        self.add_table(
                            worksheet,
                            columns,
                            value
                        )

                    else:

                        for row, item in enumerate(
                            value,
                            1
                        ):

                            worksheet.cell(
                                row=row,
                                column=1
                            ).value = (
                                format_value(item)
                            )

                else:

                    worksheet.cell(
                        row=1,
                        column=1
                    ).value = (
                        format_value(value)
                    )

                self.auto_width(
                    worksheet
                )

        # ----------------------------------------------------
        # Fallback
        # ----------------------------------------------------

        if not self.workbook.sheetnames:

            worksheet = (
                self.create_sheet(
                    "Document"
                )
            )

            self.render_dict(
                worksheet,
                self.data
            )

            self.auto_width(
                worksheet
            )

        return self.workbook

    # ========================================================
    # SAVE
    # ========================================================

    def save(
        self,
        output_path
    ):

        output_path = Path(
            output_path
        )

        output_path.parent.mkdir(
            parents=True,
            exist_ok=True
        )

        self.workbook.save(
            output_path
        )

        return output_path


# ============================================================
# PUBLIC FUNCTIONS
# ============================================================

def load_json(json_path):

    with open(
        json_path,
        "r",
        encoding="utf-8"
    ) as file:

        return json.load(file)


def generate_excel_from_json(
    json_path,
    output_path=None
):

    data = load_json(
        json_path
    )

    generator = (
        AdaptiveExcelGenerator(
            data
        )
    )

    generator.render()

    if output_path is None:

        base_dir = (
            Path(json_path).parent.parent
        )

        output_dir = (
            base_dir / "excel_output"
        )

        output_dir.mkdir(
            exist_ok=True
        )

        name = safe_filename(
            data.get(
                "document_name",
                Path(json_path).stem
            )
        )

        output_path = (
            output_dir
            / f"{name}.xlsx"
        )

    return generator.save(
        output_path
    )